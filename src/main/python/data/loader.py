from openpyxl import load_workbook
from itertools import zip_longest
from typing import Tuple, List, Set

CARDS_PER_WORKSHEET = 5
BATCH_NUMBER_COLUMN_NAME = "J"
BATCH_NUMBER_COLUMN_OFFSET = 9  # batch numbers are in col J
CAC_ENV_COL_NAME = "C"
PROXY_NUMBER_COL_NAME = "D"


# from https://docs.python.org/3.8/library/itertools.html
def grouper(iterable, n, fillvalue=None):
    "Collect data into fixed-length chunks or blocks"
    # grouper('ABCDEFG', 3, 'x') --> ABC DEF Gxx"
    args = [iter(iterable)] * n
    return zip_longest(*args, fillvalue=fillvalue)

def load_sheet_names(path_to_spreadsheet: str) -> List[str]:
    wb = load_workbook(filename=path_to_spreadsheet)
    return wb.sheetnames


def load_batch_numbers_from_inventory_file(path_to_spreadsheet: str, from_sheet: str) -> Set[str]:
    """
    Rolls through a inventory spreadsheet, returns unique batch numbers found
    :param path_to_spreadsheet:
    :param from_sheet:
    :return:
    """
    inventory_wb = load_workbook(filename=path_to_spreadsheet)

    if from_sheet not in inventory_wb.sheetnames:
        raise (Exception(f"Spreadsheet does not contain worksheet named {from_sheet}"))

    inventory_sheet = inventory_wb[from_sheet]

    all_j = inventory_sheet[BATCH_NUMBER_COLUMN_NAME]

    # get all the batch numbers from this table:
    all_batches = set([c.value for c in all_j[1:] if c.value])
    return all_batches

# a set of e.g. 5 cards, with the envelope number and proxy number
CardBatch = List[Tuple[int, int]]

def load_values_from_inventory_file(
        path_to_spreadsheet: str,
        from_sheet: str,
        target_batch: str,
        cards_per_sheet: int = CARDS_PER_WORKSHEET
) -> List[CardBatch]:
    """
    Retrieves the actual card and proxy numbers from the inventory worksheet
    :param path_to_spreadsheet:
    :param from_sheet:
    :param target_batch:
    :param cards_per_sheet:
    :return:
    """
    to_ret = []

    inventory_wb = load_workbook(filename=path_to_spreadsheet)

    if from_sheet not in inventory_wb.sheetnames:
        raise (Exception(f"Spreadsheet does not contain worksheet named {from_sheet}"))

    inventory_sheet = inventory_wb[from_sheet]

    all_j = inventory_sheet[BATCH_NUMBER_COLUMN_NAME]

    # figure out which rows have entries for target_batch
    # entries are contiguous, so we just need to find the min and max

    start_row = None  # these will be 1-indexed, just like Excel does
    stop_row = None

    for idx, r in enumerate(inventory_sheet.iter_rows(min_row=2)):  # skip the header row
        batch_num = r[BATCH_NUMBER_COLUMN_OFFSET].value
        if batch_num == target_batch:
            if start_row is None:  # we must have found the first occurrence
                start_row = idx + 2  # add one for the header, and one more to get us into 1-indexed territory
                stop_row = idx + 2 # we could have a batch with only one card in it
            if idx + 2 >= stop_row:
                stop_row = idx + 2

    if start_row is None or stop_row is None:
        raise (Exception(f"Couldn't find batch {target_batch} in {path_to_spreadsheet}"))

    target_cells = inventory_sheet[f"{CAC_ENV_COL_NAME}{start_row}:{PROXY_NUMBER_COL_NAME}{stop_row}"]

    for idx, chunk in enumerate(grouper(target_cells, cards_per_sheet)):
        # try and validate:
        is_valid = True
        for row in chunk:
            # We need to check to see that we are not in one of the None-padding entries that grouper will return if
            # len(target_cells) % cards_per_sheet != 0
            if row is not None:
                card_num, proxy_num = row
                if card_num.value is None or card_num.value == '':
                    is_valid = False
                if proxy_num.value is None or proxy_num.value == '':
                    is_valid = False

        if not is_valid:
            raise (Exception(f"Invalid data for chunk {idx} (somewhere around {chunk[0][0].row})"))

        # now extract values:
        actual_values = [(x[0].value, x[1].value) for x in chunk if x is not None]

        to_ret.append(actual_values)
    return to_ret
